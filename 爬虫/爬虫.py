import re
import requests
import json
import time
from openpyxl import load_workbook, Workbook
from requests import RequestException


def get_detail_page(html):
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.100 Safari/537.36"
        }
        cookies = {}
        response = requests.get(url=html, headers=headers, cookies=cookies)
        response.encoding = 'utf-8'
        if response.status_code == 200:
            return response.text
        return None
    except RequestException:
        print('获取详情页错误')
        time.sleep(3)
        return get_detail_page(html)

def parse_index_page(html):
    html = get_detail_page(html)
    html = html[12:-1]
    data = json.loads(html)
    id_list = []
    if data:
        for item in data:
            id_list.append(item['url'])
    return id_list

def parse_detail_page(data):
    html = get_detail_page(data)
    info = []
    # 获取电影名称
    name_pattern = re.compile('<span property="v:itemreviewed">(.*?)</span>')
    name = re.findall(name_pattern, html)
    info.append(name[0])
    # 获取评分
    score_pattern = re.compile('rating_num" property="v:average">(.*?)</strong>')
    score = re.findall(score_pattern, html)
    info.append(score[0])
    # 获取导演
    director_pattern = re.compile('rel="v:directedBy">(.*?)</a>')
    director = re.findall(director_pattern, html)
    info.append(str(director[0]))
    # 获取演员
    actor_pattern = re.compile('rel="v:starring">(.*?)</a>')
    actor = re.findall(actor_pattern, html)
    info.append(str(actor[0]))
    # 获取年份
    year_pattern = re.compile('<span class="year">\((.*?)\)</span>')
    year = re.findall(year_pattern, html)
    info.append(year[0])
    # 获取类型
    type_pattern = re.compile('property="v:genre">(.*?)</span>')
    type = re.findall(type_pattern, html)
    info.append(type[0].split(' /')[0])
    # 获取时长
    try:
        time_pattern = re.compile('property="v:runtime" content="(.*?)"')
        time = re.findall(time_pattern, html)
        info.append(time[0])
    except:
        info.append('1')
    # 获取语言
    language_pattern = re.compile('pl">语言:</span>(.*?)<br/>')
    language = re.findall(language_pattern, html)
    info.append(language[0].split(' /')[0])
    # 获取评价人数
    comment_pattern = re.compile('property="v:votes">(.*?)</span>')
    comment = re.findall(comment_pattern, html)
    info.append(comment[0])
    # 获取地区
    area_pattern = re.compile(' class="pl">制片国家/地区:</span>(.*?)<br/>')
    area = re.findall(area_pattern, html)
    info.append(area[0].split(' /')[0])
    return info


html = 'https://movie.douban.com/j/search_subjects?type=movie&tag=%E5%86%B7%E9%97%A8%E4%BD%B3%E7%89%87&sort=rank&page_limit=20&page_start='


wc = Workbook()
sheet = wc.active
sheet.title = "New"
ws = wc['New']
sheet['A1'] = 'name'
sheet['B1'] = 'score'
sheet['C1'] = 'director'
sheet['D1'] = 'actor'
sheet['E1'] = 'year'
sheet['F1'] = 'type'
sheet['G1'] = 'time'
sheet['H1'] = 'language'
sheet['I1'] = 'comment'
sheet['J1'] = 'area'
ws = wc[wc.sheetnames[0]]
wc.save('豆瓣电影.xlsx')

ti = 1
for i in range(20, 50):
    print(i)
    html1 = html+str(i*20)
    u = parse_index_page(html1)
    print(u)
    for t in u:
        time.sleep(0.5)
        b = parse_detail_page(t)
        print(b)
        ws.append(b)
        wc.save('豆瓣电影.xlsx')
        ti += 1



